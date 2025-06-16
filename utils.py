import os
import requests
import openpyxl
from openpyxl.utils import get_column_letter
import configparser

def load_or_ask_exchangerate_api_key():
    config = configparser.ConfigParser()
    if os.path.exists("config.ini"):
        config.read("config.ini")
        return config["exchangerate"]["exchangerate_API_Key"]
    else:
        API_Key = input("exchangerateのAPI_Keyを入力してください：").strip()
        config["exchangerate"] = {"exchangerate_API_Key": API_Key}
        with open("config.ini", "w") as configfile:
            config.write(configfile)
        return API_Key

def get_usd_to_jpy_rate(api_key: str) -> float:
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/latest/USD"
    response = requests.get(url)
    data = response.json()

    if data['result'] == 'success':
        rate = data['conversion_rates'].get('JPY')
        if rate is not None:
            return rate
        else:
            rate = 155;
            raise ValueError("JPYのレートが見つかりません。")
    else:
        rate = 155;
        raise Exception(f"APIエラー: {data.get('error-type')}")

def extract_price_info(items, name_key="商品名", price_key="価格", url_key="商品URL"):
    prices = [float(item[price_key]) for item in items if item.get(price_key)]
    if not prices:
        return {"name": "N/A", "min": "N/A", "max": "N/A", "url": "N/A"}
    min_index = prices.index(min(prices))
    max_index = prices.index(max(prices))
    return {
        "name": items[min_index][name_key],
        "min": min(prices),
        "max": max(prices),
        "url": items[min_index][url_key]
    }

def save_comparison_to_excel(filename, exchange_rate, rakuten_data, ebay_data,yahoo_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品比較"

    ws.append([f"為替レート (USD→JPY): {exchange_rate:.2f}"])
    headers = [
        "楽天商品名", "eBay商品名","Yahoo商品名",
        "楽天最安価格(円)", "eBay最安価格(円)","Yahoo最安価格(円)",
        "楽天最高価格(円)", "eBay最高価格(円)","Yahoo最高価格(円)",
        "楽天URL", "eBayURL","YahooURL"
    ]
    ws.append(headers)
    ws.append([
        rakuten_data['name'], ebay_data['name'],yahoo_data['name'],
        rakuten_data['min'], ebay_data['min'],yahoo_data['min'],
        rakuten_data['max'], ebay_data['max'],yahoo_data['max'],
        rakuten_data['url'], ebay_data['url'],yahoo_data['url']
    ])

    for col_num, col_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(col_title) + 2, 15)

    wb.save(filename)
    print(f"✅ 比較結果を保存しました：{filename}")

def save_all_results_to_excel(filename, rakuten_items, ebay_items,yahoo_items):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(filename)

    # 楽天シート
    if "楽天商品一覧" in wb.sheetnames:
        del wb["楽天商品一覧"]
    ws_rakuten = wb.create_sheet(title="楽天商品一覧")
    ws_rakuten.append(["商品名", "価格", "URL"])
    for item in rakuten_items:
        ws_rakuten.append([item.get("商品名", ""), item.get("価格", ""), item.get("商品URL", "")])

    # eBayシート
    if "eBay商品一覧" in wb.sheetnames:
        del wb["eBay商品一覧"]
    ws_ebay = wb.create_sheet(title="eBay商品一覧")
    ws_ebay.append(["商品名", "価格", "URL"])
    for item in ebay_items:
        ws_ebay.append([item.get("商品名", ""), item.get("価格", ""), item.get("商品URL", "")])

    # Yahooシート
    if "yahoo商品一覧" in wb.sheetnames:
        del wb["yahoo商品一覧"]
    ws_yahoo = wb.create_sheet(title="yahoo商品一覧")
    ws_yahoo.append(["商品名", "価格", "URL"])
    for item in yahoo_items:
        ws_yahoo.append([item.get("商品名", ""), item.get("価格", ""), item.get("商品URL", "")])
    wb.save(filename)

    print(f"✅ 全検索結果を同一ファイルに保存しました：{filename}")