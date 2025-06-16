import requests
import configparser
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# 楽天のconfigファイル
RAKUTEN_CONFIG_FILE = "config.ini"

# 楽天アプリIDを読み込むまたは尋ねる
def load_or_ask_rakuten_application_id():
    config = configparser.ConfigParser()

    # ファイルが存在するかチェック
    if os.path.exists("config.ini"):
        config.read("config.ini")

        # セクションとキーが存在するかチェック
        if "RAKUTEN" in config and "application_id" in config["RAKUTEN"]:
            return config["RAKUTEN"]["application_id"]

    # 入力と保存処理
    app_id = input("楽天アプリIDを入力してください：").strip()
    config["RAKUTEN"] = {"application_id": app_id}
    with open(RAKUTEN_CONFIG_FILE, "w") as configfile:
        config.write(configfile)
    print("アプリIDをconfig.iniに保存しました。")
    return app_id

# 楽天の商品を検索
def search_rakuten_items(keyword, application_id, hits=30):
    url = "https://app.rakuten.co.jp/services/api/IchibaItem/Search/20170706"
    params = {
        "format": "json",
        "keyword": keyword,
        "applicationId": application_id,
        "hits": hits,
    }

    response = requests.get(url, params=params)
    response.raise_for_status()
    data = response.json()

    items = data.get("Items", [])
    results = []

    for item in items:
        item_data = item["Item"]
        results.append({
            "商品名": item_data["itemName"],
            "価格": item_data["itemPrice"],
            "店舗名": item_data["shopName"],
            "商品URL": item_data["itemUrl"],
        })

    return results

# 楽天の商品情報をExcelに保存
def save_rakuten_items_to_excel(items, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "楽天商品情報"

    if not items:
        print("保存するデータがありません。")
        return

    headers = list(items[0].keys())
    ws.append(headers)

    for item in items:
        ws.append([item[h] for h in headers])

    # カラム幅自動調整
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
